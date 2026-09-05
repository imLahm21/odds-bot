#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build_calibration.py —— 从战绩表重算「置信度 → 真实胜率」校准表

用途：
    读本地 竞彩.xlsx 的下注战绩，按「置信度分桶」和「玩法分类」统计真实胜率/ROI；
    若存在「公允赔率」列，额外按事前 edge 分桶校准。输出可直接粘进
    rules/风控验证/reference_staking_kelly.md 第二部分观测层的 Markdown 表。

    这是【低频重跑】产物：每新增约 30 笔或每周一次即可，不需要每天跑。
    每日「这场买什么/买多少」靠 LLM 实时算的 p_市场，不依赖本脚本。
    使用说明见 build_calibration.md。

只依赖本地 竞彩.xlsx + openpyxl，不碰服务器/赔率库，只读不写表格。

用法：
    python build_calibration.py                     # 默认 3 个 sheet
    python build_calibration.py --months 3          # 只取最近 3 个月（按日期滚动窗口）
    python build_calibration.py --sheets 世界杯 20260701-20260731 20260801-20260831
    python build_calibration.py --xlsx 竞彩.xlsx     # 指定表格路径
"""

import argparse
import io
import os
import re
import sys
from datetime import datetime, timedelta

try:
    import openpyxl
except ImportError:
    sys.exit("缺少 openpyxl，请先: pip install openpyxl")

# 默认统计的 sheet（用户选定的 3 个有效窗口）
# 更早的表因模型改动过多、判读尺子已漂移，剔除以免污染校准。
DEFAULT_SHEETS = [
    "世界杯",
    "20260701-20260731",
    "20260801-20260831",
]

# 置信度分桶边界（左闭右开）：与 reference_staking_kelly.md 第二章一致
CONF_BUCKETS = [(0, 60), (60, 64), (64, 66), (66, 70), (70, 74), (74, 200)]
CONF_LABELS = ["< 60", "60–64", "64–66", "66–70", "70–74", "74+"]

# 玩法分类顺序（输出用）
PLAY_CATS = ["让球", "大小球", "双进", "胜平负/其他", "波胆", "串关"]

# 支持以后给每注补记公允赔率；当前表没有该列时自动跳过事前 edge 表。
FAIR_ODDS_HEADERS = ("公允赔率", "我认为的公允赔率", "我的公允赔率", "公平赔率")

SCORE_RE = re.compile(r"\d+\s*[:：]\s*\d+")
RESULT_SUFFIX_RE = re.compile(
    r"\s*[（(]\s*\d+\s*[:：]\s*\d+"
    r"(?:\s*\+\s*\d+\s*[:：]\s*\d+)*\s*[）)]\s*$"
)


def classify_play(play: str) -> str:
    """按投注玩法文本分类。规则与 plan 约定一致。"""
    p = str(play or "").strip()

    # 先识别串关。赛果可能写成「（2:4+1:1）」；两条让球腿也可能直接
    # 写成「巴拿马+2.5+克罗地亚-0.5」。这些都必须在玩法分类前截住。
    if "&" in p or re.search(r"\bvs\b", p, flags=re.IGNORECASE):
        return "串关"
    if len(SCORE_RE.findall(p)) >= 2 or re.search(r"[）)]\s*\+", p):
        return "串关"

    base = RESULT_SUFFIX_RE.sub("", p).strip()
    signed_lines = re.findall(r"[+-]\s*\d+(?:\.\d+)?", base)
    if len(signed_lines) >= 2:
        return "串关"
    if re.search(
        r"(?:胜|平|双进|[大小]\s*\d+(?:\.\d+)?|[+-]\s*\d+(?:\.\d+)?)"
        r"\s*\+\s*(?=[^\d])",
        base,
    ):
        return "串关"

    # 「平（1:1）」中的 1:1 是赛果，不是波胆玩法；只有明确写“波胆”
    # 或玩法本身就是一个比分时才归入波胆。
    if "波胆" in base or re.fullmatch(r"\d+\s*[:：]\s*\d+", base):
        return "波胆"
    if "双进" in base:
        return "双进"
    # 不能只搜“大/小”：例如“大连英博-0.75”是让球，不是大小球。
    if re.search(r"[大小]\s*\d+(?:\.\d+)?", base):
        return "大小球"
    if re.search(r"[-+]\s*\d+(?:\.\d+)?", base) or "受" in base:
        return "让球"
    # 平手盘在表里常写成「捷克0」「上海 0」。
    if re.search(r"0\s*$", base):
        return "让球"
    return "胜平负/其他"


def header_map(ws) -> dict:
    """表头名 → 列号（1-based）。取第一行；重名保留最左。"""
    h = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v is not None:
            h.setdefault(str(v).strip(), c)
    return h


def parse_date(v):
    """把 A 列日期解析成 datetime（兼容 datetime 对象与 '2026/7/3-1:00' 等字符串）。"""
    if isinstance(v, datetime):
        return v
    s = str(v or "").strip()
    if not s:
        return None
    # 取前面的日期部分（去掉 '-1:00' 之类的时间尾巴）
    s = re.split(r"[ \-T]", s)[0] if "/" in s else s.split(" ")[0]
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_decimal_odds(v):
    """把可选的公允赔率解析成十进制赔率；无效或不大于 1 时返回 None。"""
    try:
        value = float(v)
    except (TypeError, ValueError):
        return None
    return value if value > 1 else None


def collect_rows(xlsx_path, sheets, months):
    """收集带置信度的决策行；末项 fair_odds 可为 None。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    cutoff = None
    if months:
        cutoff = datetime.now() - timedelta(days=int(round(months * 30.4)))
    rows = []
    seen_sheets = []
    for ws in wb.worksheets:
        if ws.title not in sheets:
            continue
        seen_sheets.append(ws.title)
        h = header_map(ws)
        ci = h.get("置信度")
        oi = h.get("赔率")
        si = h.get("投入金额") or h.get("投入金额 ")
        ni = h.get("净盈亏")
        ri = h.get("结果")
        pi = h.get("投注玩法")
        di = h.get("日期")
        fi = next((h.get(name) for name in FAIR_ODDS_HEADERS if h.get(name)), None)
        if not (ci and oi and ni):
            continue
        for r in range(2, ws.max_row + 1):
            conf = ws.cell(r, ci).value
            if not isinstance(conf, (int, float)):
                continue  # 只统计带数值置信度的「决策」行
            net = ws.cell(r, ni).value
            odds = ws.cell(r, oi).value
            if net is None or odds is None:
                continue
            date = parse_date(ws.cell(r, di).value) if di else None
            if cutoff and date and date < cutoff:
                continue
            stake = ws.cell(r, si).value if si else 0
            result = ws.cell(r, ri).value if ri else ""
            play = ws.cell(r, pi).value if pi else ""
            fair_odds = parse_decimal_odds(ws.cell(r, fi).value) if fi else None
            rows.append((float(conf), float(odds), float(stake or 0), float(net),
                         str(result or "").strip(), str(play or ""), ws.title, date,
                         fair_odds))
    missing = [s for s in sheets if s not in seen_sheets]
    return rows, seen_sheets, missing


# ─── 结算状态分类（五态）─────────────────────────────────────────────────────
# 「结果」列的真实取值（实测 496 笔）：黑 255、红 177、和 27、输一半 21、赢一半 15。
# 非二元结算占 12.7% —— 必须分开计，不能一律按「净盈亏>0」二分，否则算出的
# 既不是全赢概率、也不是方向命中率，而是「盈利笔数占比」。
_WIN_FULL, _WIN_HALF, _PUSH, _LOSE_HALF, _LOSE_FULL = (
    "win", "half_win", "push", "half_lose", "lose")


def settle_state(result: str, net: float) -> str:
    """把「结果」列文本归一到五态；文本缺失时退回按净盈亏二分。"""
    s = (result or "").replace(" ", "")
    if "赢一半" in s or "赢半" in s:
        return _WIN_HALF
    if "输一半" in s or "输半" in s:
        return _LOSE_HALF
    if s in ("和", "走", "走水", "走盘"):
        return _PUSH
    if s == "红":
        return _WIN_FULL
    if s == "黑":
        return _LOSE_FULL
    # 混合串关（如「红+黑=黑」）或无文本：按净盈亏定性，不细分半赢半输
    if net > 1e-4:
        return _WIN_FULL
    if net < -1e-4:
        return _LOSE_FULL
    return _PUSH


def stat(sel):
    """给一组行算统计量。返回
    (n, 方向命中率%, ROI%, 净盈亏, 全赢率%, 走盘率%, 盈利笔数占比%)。

    ⚠️ 三个「率」口径不同，勿混用：
      · **方向命中率** = (全赢 + 半赢) / (总数 − 走盘)。走盘不计入分母——
        它是「方向没错但没赚到」，把它算成未命中会低估判断力。
        这是 p_校准 该用的口径（B1：校准「我方选中方向的正确率」）。
      · **全赢率** = 全赢 / 总数。四分之一盘的半赢不算，偏保守。
      · **盈利笔数占比** = 净盈亏>0 的笔数 / 总数（旧版口径，含半赢、排除走盘与半输）。
        仅作报表参考，**不可当概率用**。
    """
    n = len(sel)
    if n == 0:
        return (0, None, None, 0.0, None, None, None)
    states = [settle_state(x[4], x[3]) for x in sel]
    n_win = states.count(_WIN_FULL)
    n_hw = states.count(_WIN_HALF)
    n_push = states.count(_PUSH)
    denom = n - n_push                      # 方向命中率的分母剔除走盘
    hit = (100 * (n_win + n_hw) / denom) if denom else None
    stk = sum(x[2] for x in sel)
    net = sum(x[3] for x in sel)
    roi = (100 * net / stk) if stk else None
    profit_cnt = sum(1 for x in sel if x[3] > 1e-4)
    return (n, hit, roi, net,
            100 * n_win / n, 100 * n_push / n, 100 * profit_cnt / n)


def fmt_pct(v):
    return "—" if v is None else f"{v:.1f}%"


def fmt_roi(v):
    return "—" if v is None else f"{v:+.1f}%"


def build_fair_odds_markdown(rows):
    """按事前 edge 分桶，比较主观概率、实测胜率与实际 ROI。"""
    fair_rows = [x for x in rows if x[8] is not None]
    if not fair_rows:
        return []

    edge_buckets = [
        ("≤ 0%", lambda edge: edge <= 0),
        ("0–5%", lambda edge: 0 < edge < 0.05),
        ("5–10%", lambda edge: 0.05 <= edge < 0.10),
        ("10–20%", lambda edge: 0.10 <= edge < 0.20),
        ("20%+", lambda edge: edge >= 0.20),
    ]

    out = [
        "### 事前 edge 校准（有公允赔率的样本）\n",
        "| 事前 edge | 笔数 | 平均事前 edge | 平均主观胜率 | 实测胜率 | 校准偏差 | ROI | 净盈亏 |",
        "|-----------|------|---------------|--------------|---------|---------|-----|--------|",
    ]
    for label, includes in edge_buckets:
        sel = [x for x in fair_rows if includes(x[1] / x[8] - 1)]
        if not sel:
            continue
        n, wr, roi, net, _fullw, _pushr, _profit = stat(sel)
        avg_edge = 100 * sum(x[1] / x[8] - 1 for x in sel) / n
        avg_p = 100 * sum(1 / x[8] for x in sel) / n
        gap = wr - avg_p if wr is not None else None
        out.append(
            f"| {label} | {n} | {avg_edge:+.1f}% | {avg_p:.1f}% "
            f"| {fmt_pct(wr)} | {fmt_roi(gap)} | {fmt_roi(roi)} | {net:+.2f} |"
        )
    return out


def build_markdown(rows):
    """产出可直接更新到 Kelly 文档观测层的 Markdown。"""
    out = []
    out.append("### 置信度校准表（本人实测，Kelly 的 p_校准 从「方向命中率」列取）\n")
    out.append("> 口径：**方向命中率** =（全赢+半赢）/（总数−走盘），"
               "走盘不计入分母（方向没错、只是没赚到）。"
               "**全赢率**=全赢/总数（半赢不算，偏保守）。"
               "**盈利笔数**=净盈亏>0 的占比（旧口径，仅作参考，**不可当概率用**）。")
    out.append("")
    out.append("| 自评置信度 | 笔数 | 方向命中率 | 全赢率 | 走盘率 | 盈利笔数 | ROI | 净盈亏 | 判定 |")
    out.append("|-----------|------|-----------|--------|--------|---------|-----|--------|------|")
    for (lo, hi), label in zip(CONF_BUCKETS, CONF_LABELS):
        sel = [x for x in rows if lo <= x[0] < hi]
        n, wr, roi, net, fullw, pushr, profit = stat(sel)
        verdict = "样本不足" if n < 15 else (
            "✅ 盈利" if (roi is not None and roi > 0) else "❌ 亏")
        out.append(f"| {label} | {n} | {fmt_pct(wr)} | {fmt_pct(fullw)} "
                   f"| {fmt_pct(pushr)} | {fmt_pct(profit)} | {fmt_roi(roi)} "
                   f"| {net:+.2f} | {verdict} |")
    out.append("")
    out.append("### 玩法分类表现\n")
    out.append("| 玩法 | 笔数 | 命中 | ROI | 净盈亏 | 用法 |")
    out.append("|------|------|------|-----|--------|------|")
    for cat in PLAY_CATS:
        sel = [x for x in rows if classify_play(x[5]) == cat]
        n, wr, roi, net, _fullw, _pushr, _profit = stat(sel)
        if n == 0:
            use = "样本为空"
        elif cat in ("波胆", "串关"):
            use = "❌ 拉黑"
        elif roi is not None and roi > 5:
            use = "✅ 优先" if n >= 10 else "✅ 优先（小样本）"
        elif roi is not None and roi >= -5:
            use = "≈ 持平"
        else:
            use = "⚠️ 降级"
        out.append(f"| {cat} | {n} | {fmt_pct(wr)} | {fmt_roi(roi)} "
                   f"| {net:+.2f} | {use} |")
    fair_odds_out = build_fair_odds_markdown(rows)
    if fair_odds_out:
        out.append("")
        out.extend(fair_odds_out)
    return "\n".join(out)


def main():
    ap = argparse.ArgumentParser(
        description="从 竞彩.xlsx 重算置信度→真实胜率校准表（低频重跑，非每天）")
    ap.add_argument("--xlsx", default="竞彩.xlsx", help="战绩表路径（默认 竞彩.xlsx）")
    ap.add_argument("--sheets", nargs="+", default=None,
                    help="指定统计的 sheet 名（默认用户选定的 3 个）")
    ap.add_argument("--months", type=float, default=None,
                    help="只取最近 N 个月（按 A 列日期滚动窗口；缺日期的行仍计入）")
    ap.add_argument("--out", default=None,
                    help="把 Markdown 另存到文件（默认只打印到终端）")
    args = ap.parse_args()

    if not os.path.exists(args.xlsx):
        sys.exit(f"找不到战绩表: {args.xlsx}")
    sheets = args.sheets or DEFAULT_SHEETS

    rows, seen, missing = collect_rows(args.xlsx, sheets, args.months)

    # Windows 终端 GBK 常打不出部分字符，统一用 UTF-8 包一层 stdout
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass

    print(f"# 校准表（生成于 {datetime.now():%Y-%m-%d %H:%M}）\n")
    print(f"- 数据源：{args.xlsx}")
    print(f"- 统计 sheet：{', '.join(seen) if seen else '（无匹配）'}")
    if missing:
        print(f"- ⚠️ 未找到 sheet：{', '.join(missing)}")
    if args.months:
        print(f"- 滚动窗口：最近 {args.months} 个月")
    print(f"- 有效决策样本（带数值置信度）：{len(rows)} 笔")

    fair_odds_count = sum(1 for x in rows if x[8] is not None)
    fair_note = "" if fair_odds_count else "（未找到公允赔率列或有效值，本期不生成事前 edge 表）"
    print(f"- 有效公允赔率样本：{fair_odds_count} 笔{fair_note}\n")

    if not rows:
        print("（无样本，检查 sheet 名或 --months 是否过窄）")
        return

    md = build_markdown(rows)
    print(md)

    if args.out:
        with io.open(args.out, "w", encoding="utf-8") as f:
            f.write(f"# 校准表（生成于 {datetime.now():%Y-%m-%d %H:%M}）\n\n")
            f.write(f"- 数据源：{args.xlsx}\n")
            f.write(f"- 统计 sheet：{', '.join(seen) if seen else '（无匹配）'}\n")
            f.write(f"- 有效决策样本：{len(rows)} 笔\n")
            f.write(f"- 有效公允赔率样本：{fair_odds_count} 笔{fair_note}\n\n")
            f.write(md + "\n")
        print(f"\n已写入：{args.out}")


if __name__ == "__main__":
    main()
